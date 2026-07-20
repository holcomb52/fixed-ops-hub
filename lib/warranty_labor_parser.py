"""Parse warranty labor rate spreadsheets."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import BinaryIO, Dict, List, Tuple, Union

import openpyxl

from lib.warranty_labor_calc import WarrantyLaborRow

# Standard DMS / dealership export layout
DMS_HEADER_MAP = {
    "RECID": "recid",
    "RO-DATE": "ro_date",
    "ADV-NO": "advisor_no",
    "CWI-FLAG": "cwi_flag",
    "CWI": "cwi_flag",
    "SVC-OP-CODES": "op_code",
    "OP-DESC": "op_desc",
    "TECH HRS": "tech_hrs",
    "LBR COST": "lbr_cost",
    "LBR SALE": "lbr_sale",
    "LBR-GROSS": "lbr_gross",
    "ELR": "sheet_elr",
    "FIRST-NAME": "first_name",
    "LAST-NAME": "last_name",
    "STD-MK-CODE": "make_code",
    "MISC CODE": "misc_code",
    "NOTES": "notes",
}

# Attorney / FINAL customer-pay RO list layout
ATTORNEY_HEADER_MAP = {
    "RO#": "recid",
    "RO #": "recid",
    "RO-DATE": "ro_date",
    "MAKE": "make_code",
    "CWI": "cwi_flag",
    "CWI-FLAG": "cwi_flag",
    "JOB-NO": "op_code",
    "JOB #": "op_code",
    "OP-DESC": "op_desc",
    "HOURS": "tech_hrs",
    "BILL-AMT": "lbr_sale",
    "DISCOUNT": "discount",
}

# Back-compat alias used by older callers / tests
HEADER_MAP = DMS_HEADER_MAP


def _fmt_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).strftime("%m/%d/%Y")
        except ValueError:
            continue
    return text


def _float(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _normalize_header(value) -> str:
    text = str(value or "").strip().upper()
    text = text.replace("_", "-")
    while "  " in text:
        text = text.replace("  ", " ")
    return text


def list_sheet_names(source: Union[str, BytesIO, BinaryIO]) -> List[str]:
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    return wb.sheetnames


def _detect_header_row(ws) -> Tuple[int, List[str]]:
    """Find the first row that looks like a warranty labor header."""
    max_scan = min(15, ws.max_row or 1)
    for row_num in range(1, max_scan + 1):
        headers = [_normalize_header(cell.value) for cell in ws[row_num]]
        header_set = {h for h in headers if h}
        if "RECID" in header_set or "RO#" in header_set or "RO #" in header_set:
            return row_num, headers
    # Fall back to row 1 so callers still get a clear missing-column error
    return 1, [_normalize_header(cell.value) for cell in ws[1]]


def _build_col_index(headers: List[str], header_map: Dict[str, str]) -> Dict[str, int]:
    col_index: Dict[str, int] = {}
    for idx, header in enumerate(headers):
        field = header_map.get(header)
        if field and field not in col_index:
            col_index[field] = idx
    return col_index


def _detect_format(headers: List[str]) -> Tuple[str, Dict[str, int]]:
    header_set = {h for h in headers if h}
    attorney_index = _build_col_index(headers, ATTORNEY_HEADER_MAP)
    dms_index = _build_col_index(headers, DMS_HEADER_MAP)

    # Prefer attorney layout when RO# + HOURS + BILL-AMT are present
    if {"recid", "tech_hrs", "lbr_sale"}.issubset(attorney_index) and (
        "RO#" in header_set or "RO #" in header_set or "HOURS" in header_set
    ):
        return "attorney", attorney_index

    if {"recid", "tech_hrs", "lbr_sale"}.issubset(dms_index):
        return "dms", dms_index

    # Prefer whichever map got closest so the error message is useful
    if len(attorney_index) >= len(dms_index):
        return "attorney", attorney_index
    return "dms", dms_index


def _cell_value(row_cells, col_index: Dict[str, int], field: str, default=""):
    idx = col_index.get(field)
    if idx is None or idx >= len(row_cells):
        return default
    value = row_cells[idx]
    return "" if value is None else value


def _build_row(
    *,
    line_index: int,
    recid: str,
    col_index: Dict[str, int],
    row_cells,
    fmt: str,
) -> WarrantyLaborRow:
    discount = _float(_cell_value(row_cells, col_index, "discount"))
    notes = str(_cell_value(row_cells, col_index, "notes") or "").strip()
    if fmt == "attorney" and discount:
        discount_note = f"Discount: {discount:g}"
        notes = f"{notes} · {discount_note}".strip(" ·") if notes else discount_note

    op_code = str(_cell_value(row_cells, col_index, "op_code") or "").strip()
    # Attorney JOB-NO often arrives as a number — keep a clean job label
    if fmt == "attorney" and op_code.endswith(".0"):
        op_code = op_code[:-2]

    return WarrantyLaborRow(
        line_id=f"{line_index:04d}-{recid}",
        recid=recid,
        ro_date=_fmt_date(_cell_value(row_cells, col_index, "ro_date")),
        advisor_no=str(_cell_value(row_cells, col_index, "advisor_no") or "").strip(),
        cwi_flag=str(_cell_value(row_cells, col_index, "cwi_flag") or "").strip(),
        op_code=op_code,
        op_desc=str(_cell_value(row_cells, col_index, "op_desc") or "").strip(),
        tech_hrs=_float(_cell_value(row_cells, col_index, "tech_hrs")),
        lbr_cost=_float(_cell_value(row_cells, col_index, "lbr_cost")),
        lbr_sale=_float(_cell_value(row_cells, col_index, "lbr_sale")),
        lbr_gross=_float(_cell_value(row_cells, col_index, "lbr_gross")),
        sheet_elr=_float(_cell_value(row_cells, col_index, "sheet_elr")),
        first_name=str(_cell_value(row_cells, col_index, "first_name") or "").strip(),
        last_name=str(_cell_value(row_cells, col_index, "last_name") or "").strip(),
        make_code=str(_cell_value(row_cells, col_index, "make_code") or "").strip(),
        misc_code=str(_cell_value(row_cells, col_index, "misc_code") or "").strip(),
        notes=notes,
    )


def parse_warranty_labor_report(
    source: Union[str, BytesIO, BinaryIO],
    sheet_name: str | None = None,
) -> List[WarrantyLaborRow]:
    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    header_row, headers = _detect_header_row(ws)
    fmt, col_index = _detect_format(headers)

    required = {"recid", "tech_hrs", "lbr_sale"}
    if not required.issubset(col_index):
        missing = required - set(col_index)
        raise ValueError(
            "Missing required columns: "
            f"{', '.join(sorted(missing))}. "
            "Supported layouts: DMS (RECID / TECH HRS / LBR SALE) or "
            "attorney FINAL list (RO# / HOURS / BILL-AMT)."
        )

    rows: List[WarrantyLaborRow] = []
    line_index = 0
    for row_cells in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row_cells:
            continue
        recid_raw = _cell_value(row_cells, col_index, "recid")
        if recid_raw in (None, ""):
            continue

        recid = str(recid_raw).strip()
        if recid.endswith(".0"):
            recid = recid[:-2]

        rows.append(
            _build_row(
                line_index=line_index,
                recid=recid,
                col_index=col_index,
                row_cells=row_cells,
                fmt=fmt,
            )
        )
        line_index += 1

    return rows
